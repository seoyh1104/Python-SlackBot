# -------------------------------------------------------------------------------------------------#
# slackpost_rack_information.py: Rack Check Data Collection and Send a message to Slack            #
# -------------------------------------------------------------------------------------------------#
#  AUTHOR: Yuhui.Seo        2022/08/16                                                             #
# --< CHANGE HISTORY >-----------------------------------------------------------------------------#
#          Yuhui.Seo        2022/09/08 #001(Add config file)                                       #
#          Yuhui.Seo        2022/09/29 #002(Add Flask server and Slash command)                    #
#          Yuhui.Seo        2023/05/09 #003(Delete Flask server and Slash command function)        #
#          Yuhui.Seo        2023/05/22 #004(Modified slack block kit)                              #
# --< Version >------------------------------------------------------------------------------------#
#          Python version 3.10.0 (requires Python version 3.10 or higher.)                         #
# -------------------------------------------------------------------------------------------------#
# Main process                                                                                     #
# -------------------------------------------------------------------------------------------------#
import os
import socket
from datetime import datetime
from configparser import ConfigParser
from openpyxl import load_workbook         # requires: pip install openpyxl
from openpyxl.styles import Border, Side
import pymssql                             # requires: pip install pymssql
import pandas as pd                        # requires: pip install pandas
from tabulate import tabulate              # requires: pip install tabulate[widechars]
from slack_sdk import WebClient            # requires: pip install slack_sdk
from slack_sdk.errors import SlackApiError

class SystemInfo:
    """SystemInfo Class"""

    def __init__(self):
        # PC Name
        self.hostname = socket.gethostname()
        # IP Address
        self.ip_address = socket.gethostbyname(self.hostname)

    def get_hostname(self):
        return self.hostname

    def set_relative_file_path(self):
        """Chdir relative file path

        Returns:
            _type_: program_directory
        """
        program_directory = os.path.dirname(os.path.abspath(__file__))
        os.chdir(program_directory)
        return program_directory


class CommonFunc:
    def get_formatted_datetime(self, date_time, fmt):
        formats = {
            1: '%Y%m%d',                   # yyyyMMdd
            2: '%Y-%m-%d',                 # yyyy-MM-dd
            3: '%Y-%m-%d %H:%M:%S',        # yyyy-MM-dd HH:mm:ss
            4: '%Y%m%d%H%M%S',             # yyyyMMddHHmmss
            5: '%Y%m%d%H%M%S%f',           # yyyyMMddHHmmssffffff
            6: '%Y년 %m월 %d일',            # yyyy년 MM월 dd일
            7: '%Y년%m월%d일 %H시%M분%S초',  # yyyy년MM월dd일 HH시mm분ss초
            8: '%Y/%m/%d',                 # YY/MM/dd
        }
        fmt_str = formats.get(fmt, '%Y%m%d%H%M%S')

        if isinstance(date_time, str):
            date_time = datetime.strptime(date_time, '%Y%m%d%H%M%S')

        if fmt == 5 and date_time.microsecond > 0:
            return self.formatted_microsecond(date_time)

        return date_time.strftime(fmt_str)

    def formatted_microsecond(self, date_time):
        # Change the microsecond value to 4 digits
        formatted_ms = date_time.strftime('%f')[:4]
        return date_time.strftime('%Y%m%d%H%M%S') + formatted_ms

    def remove_prefix(self, string, delimiter):
        index = string.find(delimiter)
        if index != -1:
            return string[index + len(delimiter):]
        return string


class ReadConfig:
    def __init__(self):
        self.conf_file = 'config.ini'

    def load_config(self):
        if not os.path.exists(self.conf_file):
            raise FileNotFoundError(f"{self.conf_file} file does not exist.")
        else:
            config = ConfigParser()
            config.read(self.conf_file, encoding='utf-8')
            print(f"Loaded config: {self.conf_file}")
            return config


class MSSQLController:
    def __init__(self, config):
        self.conn = None
        self.cursor = None
        self.sql_file = config.get('files', 'sql')
        self.server = config.get('db_config', 'server')
        self.database = config.get('db_config', 'database')
        self.user = config.get('db_config', 'username')
        self.password = config.get('db_config', 'password')
        self.charset = config.get('db_config', 'charset')

    def __connect__(self):
        try:
            self.conn = pymssql.connect(
                host=self.server,
                database=self.database,
                user=self.user,
                password=self.password,
                charset=self.charset
            )
            if self.conn is not None:
                print("MSSQL Connected")
                self.cursor = self.conn.cursor(as_dict=True)
        except Exception as error:
            print(f"Error while connecting to MSSQL: {error}")
            raise error

    def __disconnect__(self):
        try:
            if self.conn is not None and self.conn.connected:
                self.conn.close()
                print("MSSQL Connection Closed")
        except Exception as error:
            print(f"Error while closing database connection: {error}")

    def execute_query(self):
        # Read SQL file
        query = self.read_sql()
        try:
            # Connect to database
            self.__connect__()

            # Execute SQL
            self.cursor.execute(query)
            print("Query executed successfully")
            # self.conn.commit()

            data_frame = pd.DataFrame(self.cursor.fetchall())

            # Close database connection
            self.__disconnect__()

            return data_frame

        except Exception as error:
            print(f"Error while executing the query: {error}")

    def read_sql(self):
        with open(self.sql_file, 'r', encoding='utf-8') as file:
            return file.read()


# TODO: 1. 엑셀 인쇄 페이지 자동설정
class CreateFile(CommonFunc):
    def __init__(self, now, config):
        self.file_directory = config.get('files', 'file_directory')
        self.file_name = config.get('files', 'file_name')
        self.sheet_title = config.get('files', 'sheet_title')
        self.tab_color_blue = config.get('files', 'tab_color_blue')
        self.border_color = config.get('files', 'border_color')
        self.now = now

    def exists_dir(self):
        current_datetime = self.get_formatted_datetime(self.now, 4)
        if not os.path.exists(self.file_directory):
            os.mkdir(self.file_directory)
        return current_datetime, self.file_directory, self.file_name

    def save_filepath(self):
        current_datetime, directory, filename = self.exists_dir()
        filepath = directory + filename + '_' + current_datetime
        return filepath

    def save_csv(self, data_frame):
        csv_file = self.save_filepath() + '.csv'
        data_frame.to_csv(csv_file, header=True, index=False, encoding='utf-8')
        print('Csv file saved successfully : ' + csv_file)
        return csv_file

    def csv_to_excel(self, data_frame):
        r_csv = pd.read_csv(self.save_csv(data_frame))
        xlsx_file = self.save_filepath() + '.xlsx'
        save_xlsx = pd.ExcelWriter(xlsx_file)
        r_csv.to_excel(save_xlsx, index=False)  # Convert to xlsx file
        save_xlsx.save() # Save as xlsx file
        print('Excel file saved successfully : ' + xlsx_file)
        return save_xlsx

    def save_excel(self, data_frame):
        work_book = load_workbook(self.csv_to_excel(data_frame))
        work_sheet = work_book.active

        self.set_sheetdata(work_sheet)
        self.set_column_size(work_sheet)
        self.set_color_border(work_sheet)

        xlsx_file = self.save_filepath() + '.xlsx'
        work_book.save(xlsx_file)
        work_book.close()
        return xlsx_file

    def set_sheetdata(self, work_sheet):
        current_datetime = self.get_formatted_datetime(self.now, 6)
        # Change sheet name
        work_sheet.title = self.sheet_title + '_' + current_datetime
        # Change the color of sheet tabs
        work_sheet.sheet_properties.tabColor = self.tab_color_blue

        # Display in number format
        column = 'P' # P열: 재고관리코드
        for row in range(2, work_sheet.max_row + 1):
            work_sheet[column + str(row)].number_format = '0'
        # Delete column
        work_sheet.delete_cols(21)  # 21번째 열: RANK 삭제
        return work_sheet

    def set_column_size(self, work_sheet):
        work_sheet.column_dimensions['A'].width = 4  # ✔
        work_sheet.column_dimensions['B'].width = 10 # 랙ID
        work_sheet.column_dimensions['C'].width = 3  # s
        work_sheet.column_dimensions['D'].width = 3  # z
        work_sheet.column_dimensions['E'].width = 3  # x
        work_sheet.column_dimensions['F'].width = 3  # y
        work_sheet.column_dimensions['G'].width = 9  # 상태
        work_sheet.column_dimensions['H'].width = 12 # 입고자명
        work_sheet.column_dimensions['I'].width = 18 # 마지막입고일
        work_sheet.column_dimensions['J'].width = 11 # 마지막작업
        work_sheet.column_dimensions['K'].width = 18 # 마지막작업일
        work_sheet.column_dimensions['L'].width = 12 # 마지막작업자
        work_sheet.column_dimensions['O'].width = 34 # 상품명
        work_sheet.column_dimensions['P'].width = 14 # 재고관리코드
        work_sheet.column_dimensions['R'].width = 5  # 수량
        work_sheet.column_dimensions['S'].width = 9  # 상태코드

        return work_sheet

    def set_color_border(self, work_sheet):
        for col in range(work_sheet.max_column):
            for row in range(work_sheet.max_row):
                cell = work_sheet.cell(row=row + 1, column=col + 1)
                cell.border = self.make_color_border()
        return work_sheet

    def make_color_border(self):
        color = self.border_color
        border = Border(left=Side(style='thin', color=color),
                        right=Side(style='thin', color=color),
                        top=Side(style='thin', color=color),
                        bottom=Side(style='thin', color=color))
        return border


# Rack Bot
#
# Rack Check List
# PCNAME, YYYY-MM-DD HH:mm:ss
# 현재 랙정상화 대상은 XX건 입니다.
# ---------------------------------------------------------
# 랙ID       s    z   x   y  상태
# XXXXXXXXX  XX  XX  XX  XX  이중입고
# XXXXXXXXX  XX  XX  58  XX  재고확인
# ---------------------------------------------------------
# Excel file
# ---------------------------------------------------------
class SlackAPI(CommonFunc):

    def __init__(self, config, sys_info, now):
        token = config.get('slack', 'bot_token')
        self.client = WebClient(token)
        self.hostname = sys_info.get_hostname()
        self.datetime = self.get_formatted_datetime(now, 3)
        self.channel_id = config.get('slack', 'channel_id')
        self.slack_msg = config.get('slack', 'slack_msg')
        self.file_name = None

    def set_data(self, data_frame):
        if not data_frame.empty:
            data_frame_cnt = len(data_frame)
            self.slack_msg = '현재 랙정상화 대상은 ' + str(data_frame_cnt) + '건 입니다.'

            data_frame = data_frame[['랙ID', 's', 'z', 'x', 'y', '상태']]
            # data_frame.index = data_frame.index+1
            data = tabulate(data_frame, tablefmt='plain', showindex=False, headers='keys')
        return data

    def post_message(self):
        try:
            self.client.chat_postMessage(
                channel=self.channel_id,
                text=self.slack_msg, # Notification message
                blocks=[
                    {
                        "type": "header",
                        "text": {
                            "type": "plain_text",
                            "text": "Rack Check List"
                        }
                    },
                    {
                        "type": "context",
                        "elements": [
                            {
                                "type": "plain_text",
                                "text": self.hostname + ", " + self.datetime
                            }
                        ]
                    },
                    {
                        "type": "section",
                        "text": {
                            "type": "mrkdwn",
                            "text": self.slack_msg
                        }
                    }
                ]
            )
        except SlackApiError as error:
            print(error.response['error'])

    def post_files_upload(self, df_to_markdown, file):
        self.file_name = self.remove_prefix(file, '/')
        try:
            self.client.chat_postMessage(
                channel=self.channel_id,
                text='랙정상 리스트입니다.', # Notification message
                blocks=[
                    {
                        "type": "header",
                        "text": {
                            "type": "plain_text",
                            "text": "Rack Check List"
                        }
                    },
                    {
                        "type": "context",
                        "elements": [
                            {
                                "type": "plain_text",
                                "text": self.hostname + ", " + self.datetime
                            }
                        ]
                    },
                    {
                        "type": "section",
                        "text": {
                            "type": "mrkdwn",
                            "text": self.slack_msg
                            + "```" + df_to_markdown + "```"
                        }
                    }
                #     {
                #         "type": "divider"
                #     }
                # ],
                # attachments=[
                #     {
                #         "fields": [
                #             {
                #                 "short": True,
                #                 "value": "```" + df_to_markdown + "```"
                #             }
                #         ],
                #         "color": "#dddddd",
                #         "mrkdwn_in": ["title", "fields"],
                #         "footer": "자동창고 랙 정상화 리스트"
                #         # "ts": create time
                #     }
                ]
            )
            self.client.files_upload(
                channels=self.channel_id,
                file=file,
                filename=self.file_name, # File name when downloaded (requires extension)
                filetype='.xlsx',
                title=self.file_name     # name of attached file
                # initial_comment= 'default:이 파일을 업로드 했습니다.'
            )
        except SlackApiError as error:
            print(error.response['error'])


def main():
    """
    Main function
    """
    sys_info = SystemInfo()
    sys_info.set_relative_file_path()
    config = ReadConfig.load_config(ReadConfig())
    now = datetime.now()

    mssql_conn = MSSQLController(config)
    data_frame = mssql_conn.execute_query()

    slack = SlackAPI(config, sys_info, now)

    # 1. Rack Check
    if not data_frame.empty:
        # 2. If there is a rack list, create csv and excel files
        file = CreateFile.save_excel(CreateFile(now, config), data_frame)
        # 2-1. Post Excel file and list to Slack
        slack.post_files_upload(slack.set_data(data_frame), file)
    else:
        # 3. If there is no rack list, post a message to Slack
        slack.post_message()


if __name__ == '__main__':
    main()

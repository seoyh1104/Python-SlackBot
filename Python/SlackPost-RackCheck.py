#--------------------------------------------------------------------------------------------------#
# SlackPost-RackCheck.py: Rack Check Data Collection and Send a message to Slack                   #
#--------------------------------------------------------------------------------------------------#
#  AUTHOR: Yuhui.Seo        2022/08/16                                                             #
#--< CHANGE HISTORY >------------------------------------------------------------------------------#
#          Yuhui.Seo        2022/09/08 #001(Add config file)                                       #
#--< Version >-------------------------------------------------------------------------------------#
#          Python version 3.10.0 (requires Python version 3.10 or higher.)                         #
#--------------------------------------------------------------------------------------------------#

#--------------------------------------------------------------------------------------------------#
# Main process                                                                                     #
#--------------------------------------------------------------------------------------------------#
import pymssql                             # requires: pip install pymssql
import pandas as pd                        # requires: pip install pandas
import dataframe_image as dfi              # requires: pip install dataframe_image
from tabulate import tabulate              # requires: pip install tabulate[widechars]
from slack_sdk import WebClient            # requires: pip install slack_sdk
from slack_sdk.errors import SlackApiError
from openpyxl import load_workbook         # requires: pip install openpyxl
from openpyxl.styles import Border, Side
from datetime import datetime
import os
import socket
import configparser
#--------------------------------------------------------------------------------------------------#
class SystemInfo:
    def system_info():
        hostname = socket.gethostname() # PC명
        # ip = socket.gethostbyname(hostname) #IP주소
        return hostname
    
    def get_current_datetime(now, format):
        match format:
            case 1:
                format = '%Y%m%d%H%M%S' # yyyyMMddHHmmss
            case 2:
                format = '%Y-%m-%d %H:%M:%S' # yyyy-MM-dd HH:mm:ss
            case 3:
                format = '%Y년%m월%d일 %H시%M분%S초' # yyyy년MM월dd일HH시mm분ss초
            case _:
                format = '%Y%m%d%H%M%S'
        current_datetime = now.strftime(format)
        return current_datetime


class CommonFunc:
    def str_slicing(str):
        index = str.rfind('/')
        str = str[index+1:]
        return str


class ReadConfig:
    def __init__(self):
        self.conf_file = 'config.ini'
    
    def load_config(self):
        if os.path.exists(self.conf_file) == False:
            raise Exception('%s file does not exist. \n' % self.conf_file)
        else: 
            config = configparser.ConfigParser()
            config.read(self.conf_file, encoding = 'utf-8')
            print("Load Config : %s" % self.conf_file)
            return config


class ReadSql:
    def read_sql(config):
        FILE_NAME = config.get('FILES', 'sql')
        f = open(FILE_NAME, 'r', encoding='utf-8')
        query = f.readlines()
        sql = ''.join(query)
        f.close()
        return sql


class MssqlController:
    def __init__(self, config):
        self.sql = ReadSql.read_sql(config)
        self.db_config = config['DB_CONFIG']
        
    def __connect__(self):
        try:
            self.conn = pymssql.connect(server = self.db_config['server'],
                                        database = self.db_config['database'],
                                        user = self.db_config['username'], 
                                        password = self.db_config['password'],
                                        charset = self.db_config['charset'])
            self.cur = self.conn.cursor(as_dict = True)
        except Exception as e:
            print('DB not connected: ', e)
            raise
            
    def execute(self):
        self.__connect__()
        self.cur.execute(self.sql)
        df = pd.DataFrame(self.fetch())
        self.__disconnect__()
        return df
    
    def fetch(self):
        return self.cur.fetchall()
    
    def __disconnect__(self):
        self.conn.close()


# TODO: 1. 엑셀 인쇄 페이지 자동설정
class CreateFile:
    def __init__(self, now, config):
        self.files_config = config['FILES']
        self.CSV = self.files_config['CSV']
        self.XLSX = self.files_config['XLSX']
        self.FILE_DIRECTORY = self.files_config['file_directory']
        self.FILE_NAME = self.files_config['file_name']
        self.now = now
    
    def exists_dir(self):
        current_datetime = SystemInfo.get_current_datetime(self.now, 1)
        if not os.path.exists(self.file_directory):
            os.mkdir(self.file_directory)
        return current_datetime, self.file_directory, self.file_name

    def save_filepath(self):
        current_datetime, directory, filename = self.exists_dir()
        filepath = directory + filename + '_' + current_datetime
        return filepath

    def save_csv(self, df):
        csv_file = self.save_filepath() + self.CSV
        df.to_csv(csv_file, header = True, index = False, encoding = 'utf-8')
        print('Csv file saved successfully : ' + csv_file)
        return csv_file

    def csv_to_excel(self, df):
        r_csv = pd.read_csv(self.save_csv(df))
        xlsx_file = self.save_filepath() + self.XLSX
        save_xlsx = pd.ExcelWriter(xlsx_file)
        r_csv.to_excel(save_xlsx, index = False) # xlsx 파일로 변환
        save_xlsx.save() #xlsx 파일로 저장
        print('Excel file saved successfully : ' + xlsx_file)
        return save_xlsx

    def save_excel(self, df):
        wb = load_workbook(self.csv_to_excel(df))
        ws = wb.active
        
        self.set_sheetdata(ws)
        self.set_column_size(ws)
        self.set_color_border(ws)
        
        xlsx_file = self.save_filepath() + '.xlsx'
        wb.save(xlsx_file)
        wb.close()
        return xlsx_file

    def set_sheetdata(self, ws):
        current_datetime = SystemInfo.get_current_datetime(self.now, 3)
        # 시트명 변경
        ws.title = self.files_config['sheet_title'] + '_' + current_datetime
        # 시트탭 색변경
        ws.sheet_properties.tabColor = self.files_config['TAB_COLOR_BLUE']
        # ws.sheet_properties.tabColor = TAB_COLOR_PINK
        # 숫자 형식으로 표시
        column = 'P' # P열:재고관리코드
        for row in range(2, ws.max_row + 1):
            ws[column + str(row)].number_format = '0'
        # 열 삭제
        ws.delete_cols(21) # 21번째 열:RANK 삭제
        return ws

    def set_column_size(self, ws):
        ws.column_dimensions['A'].width = 4  # ✔
        ws.column_dimensions['B'].width = 10 # 랙ID
        ws.column_dimensions['G'].width = 9  # 상태
        ws.column_dimensions['H'].width = 9  # 입고자명
        ws.column_dimensions['J'].width = 11 # 마지막작업
        ws.column_dimensions['K'].width = 11 # 마지막작업일
        ws.column_dimensions['L'].width = 14 # 마지막작업자
        ws.column_dimensions['O'].width = ws.column_dimensions['O'].width + 5 # 상품명
        ws.column_dimensions['S'].width = 9  # 상태코드
        return ws

    def set_color_border(self, ws):
        for col in range(ws.max_column):
            for row in range(ws.max_row):
                cell = ws.cell(row = row + 1, column = col + 1)
                cell.border = self.make_color_border()
        return ws

    def make_color_border(self):
        color = self.files_config['BORDER_COLOR']
        border = Border(left = Side(style='thin', color=color),
                        right = Side(style='thin', color=color),
                        top = Side(style='thin', color=color),
                        bottom = Side(style='thin', color=color))
        return border


# Rack Bot
#
# Rack Check List
# PCNAME, YYYY-MM-DD HH:mm:ss
# 현재 랙정상화 대상은 XX건 입니다.
# ---------------------------------------------------------
# 랙ID       s    z   x   y  상태
# XXXXXXXXX  XX  XX  XX  XX  이중입고
# XXXXXXXXX  XX  XX  58  XX  재고확인
# ---------------------------------------------------------
# Image file
# Excel file
# ---------------------------------------------------------
class SlackAPI:

    def __init__(self, token, config, now):
        # 슬랙 클라이언트 인스턴스 생성
        self.client = WebClient(token)
        self.hostname = SystemInfo.system_info()
        self.datetime = SystemInfo.get_current_datetime(now, 2)
        files_config = config['FILES']
        slack_config = config['SLACK']
        self.XLSX = files_config['XLSX']
        self.channel_id = slack_config['channel_id']
        
    def post_Message(self, msg):
        try:
            response= self.client.chat_postMessage(
            channel= self.channel_id,
            text = '랙정상 리스트입니다.', # Slack 전송시 알람 메세지
            blocks=[
                {
                    "type": "header",
                    "text": {
                        "type": "plain_text",
                        "text": "Rack Check List"
                    }
                },
                {
                    "type": "context",
                    "elements": [
                        {
                            "type": "plain_text",
                            "text": self.hostname + ", " + self.datetime
                        }
                    ]
                },
                {
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": msg
                    }
                }
            ]
        )
        except SlackApiError as e:
            print(e.response['error'])

    def post_files_upload(self, msg, result, file):
        file_name = CommonFunc.str_slicing(file)
        try:
            response_msg = self.client.chat_postMessage(
                channel= self.channel_id,
                text = '랙정상 리스트입니다.', # Slack 전송시 알람 메세지
                blocks=[
                    {
                        "type": "header",
                        "text": {
                            "type": "plain_text",
                            "text": "Rack Check List"
                        }
                    },
                    {
                        "type": "context",
                        "elements": [
                            {
                                "type": "plain_text",
                                "text": self.hostname + ", " + self.datetime
                            }
                        ]
                    },
                    {
                        "type": "section",
                        "text": {
                            "type": "mrkdwn",
                            "text": msg
                        }
                    }
                ]
            #     attachments=[
            #         {
            #             "fallback": "⚠️요청이 실패했습니다.", # 요청 실패시 메세지
            #             "color": "#FF9500",
            #             "pretext": msg,
            #             "fields": [
            #                 {
            #                     # "title": "랙ID               s    z    x    y   상태",
            #                     "value": result
            #                 }
            #             ],
            #         }
            #     ]
            # )
            # response_xlsx1= self.client.files_upload(
            #     channels= self.channel_id,
            #     file= 'image.png',
            #     filename= 'image.png',
            #     filetype= 'png',
            #     title= 'image.png',
            #     # initial_comment= 'default:이 이미지를 업로드 했습니다.'
            )
            response_xlsx3= self.client.files_upload(
                channels= self.channel_id,
                file= file,
                filename= file_name, # 다운로드했을때의 파일명(확장자까지 설정 필요)
                filetype= self.XLSX,
                title= file_name # 첨부파일의 파일명
                # initial_comment= 'default:이 파일을 업로드 했습니다.'
            )
        except SlackApiError as e:
            print(e.response['error'])


class SlackPayload:
    def set_msg(df):
        if df.empty:
            msg = '현재 랙정상화 대상은 없습니다! :tada:'
        else:
            df_cnt = len(df)
            msg = '현재 랙정상화 대상은 ' + str(df_cnt) + '건 입니다.'
        return msg
    
    def set_result(df):
        df = df[['랙ID', 's', 'z', 'x', 'y', '상태']]
        df.index = df.index+1
        data = tabulate(df, tablefmt="plain", showindex=False) # headers='keys'
        # Create image file
        dfi.export(df, 'image.png', max_cols = -1, max_rows = -1)
        return data


#--------------------------------------------------------------------------------------------------#
# Code Entry                                                                                       #
#--------------------------------------------------------------------------------------------------#
def main():
    config = ReadConfig.load_config(ReadConfig())
    now = datetime.now()
    SLACK_TOKEN = config.get('SLACK', 'SLACK_TOKEN') # config['SLACK']['SLACK_TOKEN'] 결과동일
    slack = SlackAPI(SLACK_TOKEN, config, now)
    
    # 1.랙정상화 대상 조회 # TODO : Slack에서 /커맨드로 실행
    df = MssqlController.execute(MssqlController(config))
    if df.empty:
        # 2.랙정상화 대상이 없다면 Slack 전송
        slack.post_Message(SlackPayload.set_msg(df)) 
    else:
        # 3.랙정상화 대상이 있다면 csv, excel file 생성
        file = CreateFile.save_excel(CreateFile(now, config), df)
        # 4. Slack 전송, Excel file 전송
        slack.post_files_upload(SlackPayload.set_msg(df), SlackPayload.set_result(df), file)
    
if __name__ == "__main__":
    main()